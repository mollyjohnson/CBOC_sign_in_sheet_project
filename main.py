#!/usr/bin/python3 

# Author: Molly Johnson
# Date: 3/24/21
# Description: Creates cboc signin excel file for
# checking in clinics for every month in one year. Will adjust 
# for the days of the month, weekends, and federal holidays

# function comment template
####################################################################
### Function Title:
### Arguments:
### Returns:
### Description: 
###################################################################

# import necessary libraries 
from winreg import ExpandEnvironmentStrings
from openpyxl import Workbook
from datetime import datetime
import calendar
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import os
import federalHolidayCalculator
import readFile

# create "constants"
FZ = "Frozen"
TCH = "Tech"
TOA = "Time of Arrival"
SMP = "SMP Check"
CBOC = "_cboc_signin_sheets"
USER_INPUT_LENGTH = 4
MIN_YEAR = 2021
MID_DATE = 15
START_MONTH = 1
END_MONTH = 12
CBOC_COL_WIDTH = 12.5
WEEKEND_AND_HOLIDAY_COL_WIDTH = 2.5
HEADER_ROW_HEIGHT = 40
CBOC_ROW_HEIGHT = 18
DATE_ROW_HEIGHT = 18
TECH_TOA_ROW_HEIGHT = 27
CBOC_NAME_AND_FROZEN_ROW_HEIGHT = 18 
SPACER_ROW_HEIGHT = 6
SMP_ROW_HEIGHT = 18
HEADER_ROW = 1
HEADER_AND_LABELS_COL = 1
CBOC_COL = 1
CBOC_ROW = 2
DATE_ROW = 3
TECH_TOA_ROW = 4
CBOC_NAME_START_ROW = 5
NUM_FIXED_ROWS =4

# set "constant" cell border values
THIN = Side(border_style = "thin", color = "000000")
DOUBLE = Side(border_style = "medium", color = "000000")
THICK = Side(border_style = "thick", color = "001C54")
CBOC_NAME_BORDER = Border(top = THICK , left = THICK, right = THICK, bottom = THICK) 
TIMES_NEW_ROMAN_FONT = Font(name = 'Times New Roman', size = 10, bold = True)
DATE_BORDER = Border(left = THICK, right = THICK, bottom = THICK)
BIG_SPACE_BORDER = Border(left = THICK, right = THICK, bottom = DOUBLE)
SPACER_BORDER = Border(left = THICK, right = THICK)
CBOC_NAME_ONLY_BORDER = Border(top = DOUBLE, left = THICK, right = THICK, bottom = THIN)
FROZEN_ONLY_BORDER = Border(left = THICK, right = THICK, bottom = DOUBLE)
FROZEN_SMP_BORDER = Border(left = THICK, right = THICK, top = THIN, bottom = THIN)
SMP_ONLY_BORDER = Border(left = THICK, right = THICK, bottom = DOUBLE)
BOTTOM_ROW_BORDER = Border(top = THICK)
DATE_FONT = Font(name = 'Calibri', size = 11, bold = True)
TECH_FONT = Font(name = 'Calibri', size = 9)
TOA_FONT = Font(name = 'Calibri', size = 5)
DATE_BORDER_LEFT = Border(top = THICK, left = THICK, bottom = THICK)
DATE_BORDER_RIGHT = Border(top = THICK, right = THICK, bottom = THICK)
TECH_INFO_BORDER_LEFT = Border(left = THICK, right = THIN, bottom = DOUBLE)
TECH_INFO_BORDER_RIGHT = Border(right = THICK, bottom = DOUBLE)
SIGNATURE_BORDER_TOP_LEFT = Border(top = DOUBLE, left = THICK, right = THIN, bottom = THIN)
SIGNATURE_BORDER_BOTTOM_LEFT = Border(left = THICK, right = THIN, bottom = DOUBLE)
SIGNATURE_BORDER_TOP_RIGHT = Border(top = DOUBLE, right = THICK, bottom = THIN)
SIGNATURE_BORDER_BOTTOM_RIGHT = Border(right = THICK, left = THIN, bottom = DOUBLE)
SPACER_BORDER_LEFT = Border(right = THIN, left = THICK)
SPACER_BORDER_RIGHT = Border(right = THICK, left = THIN)
WEEKEND_AND_HOLIDAY_FILL_COLOR = PatternFill(fill_type = "solid", start_color = "BFBFBF", end_color = "BFBFBF")

####################################################################
### Function Title:
### Arguments:
### Returns:
### Description: 
####################################################################
#def createSignatureBorders(ws, endCol, noSMPCBOCs, smpCBOCs):


####################################################################
### Function Title:
### Arguments:
### Returns:
### Description: 
####################################################################
def createHeader(ws, startRow, startCol, endRow, endCol, startDateObj):
    # create header border formatting
    headerBorderLeft = Border(top = THICK , left = THICK, right = None, bottom = THICK) 
    headerBorderRight = Border(top = THICK , left = None, right = THICK, bottom = THICK)  
    headerBorderMid = Border(top = THICK, left = None, right = None, bottom = THICK)
    
    # font values
    headerFont = Font(name = 'Times New Roman', size = 28, bold = True)    
    
    ws.cell(row = startRow, column = startCol).alignment = Alignment(vertical = 'bottom')
    ws.cell(row = startRow, column = startCol).alignment = Alignment(horizontal = 'center')
    ws.cell(row = startRow, column = startCol).font = headerFont
    
    # set border at far left and far right of header merged cells
    ws.cell(row = startRow, column = startCol).border = headerBorderLeft
    ws.cell(row = endRow, column = endCol).border = headerBorderRight
    
    # set border at middle header merged cells
    for row in ws.iter_rows(min_row = startRow, max_row = endRow, min_col = (startCol + 1), max_col = (endCol - 1)):
        for cell in row:
            cell.border = headerBorderMid 
            
    # create header and merge cells A1 through AE1
    ws.cell(row = startRow, column = startCol).value = ("Month/Year: " + (str(calendar.month_name[startDateObj.month]) +
        " " + str(startDateObj.year)).upper())
    data = ws.cell(row = startRow, column = startCol).value
    ws.merge_cells(start_row = startRow, start_column = startCol, end_row = endRow, end_column = endCol)
    ws.cell(row = startRow, column = startCol).value = data

####################################################################
### Function Title: isValidUserInput()
### Arguments: user input (a 4-digit year in string format)
### Returns: boolean
### Description: checks if the user entered a 4-digit int, and if it's
### a year greater than or equal to the current year (2021). returns
### true if valid, false otherwise
####################################################################
def isValidUserInput(userInput):
    # check that length of user input string is correct
    if(len(userInput) != USER_INPUT_LENGTH):
        return False
    # is / or -, and last 2 chars are digits.
    i = 0
    while(i < USER_INPUT_LENGTH):
        if(userInput[i].isdigit() == False):
            return False
        i += 1
    # check that year is no earlier than 2021 (year this program written)
    if (int(userInput) < MIN_YEAR):
        return False
    # otherwise met all requirements, return true
    return True

####################################################################
### Function Title: mergeDate()
### Arguments: worksheet, start row, start column
### Returns: nothing
### Description: merges two date info cells
####################################################################
def mergeDateInfo(ws, startRow, startCol):
    data = ws.cell(row = startRow, column = startCol).value
    ws.merge_cells(start_row = startRow, start_column = startCol, end_row = startRow, end_column = startCol + 1)
    ws.cell(row = startRow, column = startCol).value = data

####################################################################
### Function Title: isHoliday()
### Arguments: holiday dates list, date num (date of the month)
### Returns: boolean
### Description: checks if a specific date is in the previously
### calculated holiday dates list. if yes returns true otherwise false
###################################################################
def isHoliday(holidayDates, dateNum):
    # if the date is not in the holiday dates list or if the list of
    # holiday dates is empty, the date isn't a holiday
    if((holidayDates == None) or (dateNum not in holidayDates)):
        return False
    return True

####################################################################
### Function Title: isWeekend()
### Arguments: day name
### Returns: boolean
### Description: checks if the day name passed in is a weekend day
### name. if yes returns true, otherwise returns false
####################################################################
def isWeekend(dayName):
    if(dayName == "SAT" or dayName == "SUN"):
        return True
    return False

####################################################################
### Function Title: setTechInfo()
### Arguments: worksheet, current column, current row
### Returns: nothing
### Description: adjusts the font/border for the tech and time of
### arrival cells
####################################################################
def setTechInfo(ws, curCol, curRow):
    ws.cell(row = curRow + 2, column = curCol).value = TCH
    ws.cell(row = curRow + 2, column = curCol).font = TECH_FONT
    ws.cell(row = curRow + 2, column = curCol).border = TECH_INFO_BORDER_LEFT
    

    ws.cell(row = curRow + 2, column = curCol + 1).value = TOA
    ws.cell(row = curRow + 2, column = curCol + 1).font = TOA_FONT
    ws.cell(row = curRow + 2, column = curCol + 1).border = TECH_INFO_BORDER_RIGHT
    ws.cell(row = curRow + 2, column = curCol + 1).alignment = Alignment(wrap_text=True)

####################################################################
### Function Title: setDateInfo()
### Arguments: worksheet, current column, the day of the week number (dayDate),
### the day of the week name, the date number (i.e. date of the month), current row
### Returns: nothing
### Description: sets the font/alignment/border for the day name and date of month
####################################################################
def setDateInfo(ws, curCol, dayDate, dayName, dateNum, curRow):
    # set day name
    ws.cell(row = curRow, column = curCol).value = dayName
    mergeDateInfo(ws, curRow, curCol)
    ws.cell(row = curRow, column = curCol).font = DATE_FONT
    ws.cell(row = curRow, column = curCol).alignment = Alignment(horizontal='center')
    ws.cell(row = curRow, column = curCol).border = DATE_BORDER_LEFT
    ws.cell(row = curRow, column = curCol + 1).border = DATE_BORDER_RIGHT

    # set date num
    ws.cell(row = curRow + 1, column = curCol).value = dateNum
    mergeDateInfo(ws, curRow + 1, curCol)
    ws.cell(row = curRow + 1, column = curCol).font = DATE_FONT
    ws.cell(row = curRow + 1, column = curCol).alignment = Alignment(horizontal='center')
    ws.cell(row = curRow + 1, column = curCol).border = DATE_BORDER_LEFT
    ws.cell(row = curRow + 1, column = curCol + 1).border = DATE_BORDER_RIGHT

####################################################################
### Function Title: setWeekendStyle()
### Arguments: worksheet, the current col, the current row
### Returns: nothing
### Description: adjusts the width of the columns to narrower, grays out
### all cells in the column, and places x's in the signature/time areas,
### since no CBOC's are received on weekends or federal holidays
####################################################################
def setFixedWeekendAndHolStyle(ws, curCol, curRow, endNonSmpRows, endRow):
    # adjust fill color and col width for all rows in the column
    while(curRow <= NUM_FIXED_ROWS):
        ws.cell(row = curRow, column = curCol).fill = WEEKEND_AND_HOLIDAY_FILL_COLOR      
        ws.cell(row = curRow, column = curCol + 1).fill = WEEKEND_AND_HOLIDAY_FILL_COLOR      
        ws.column_dimensions[get_column_letter(curCol)].width = WEEKEND_AND_HOLIDAY_COL_WIDTH
        ws.column_dimensions[get_column_letter(curCol + 1)].width = WEEKEND_AND_HOLIDAY_COL_WIDTH
        curRow += 1

    if(endNonSmpRows > endRow):
        endNonSmpRows = endRow

    # adjust fill color and col width for all rows in the column up to the last row 
    # for non smp rows:
    nonSMPspacerRow = curRow + 2
    while(curRow <= endNonSmpRows):
        ws.cell(row = curRow, column = curCol).fill = WEEKEND_AND_HOLIDAY_FILL_COLOR      
        ws.cell(row = curRow, column = curCol + 1).fill = WEEKEND_AND_HOLIDAY_FILL_COLOR      

        if(nonSMPspacerRow != curRow):
            ws.cell(row = curRow, column = curCol).value = "X"
            ws.cell(row = curRow, column = curCol + 1).value = "X"
            ws.cell(row = curRow, column = curCol).font = Font(name = 'Calibri', size = 15, bold = True)
            ws.cell(row = curRow, column = curCol + 1).font = Font(name = 'Calibri', size = 15, bold = True)
            nonSMPspacerRow += 3
        curRow += 1

    # for smp rows:
    SMPspacerRow = curRow + 3
    while(curRow <= endRow):
        ws.cell(row = curRow, column = curCol).fill = WEEKEND_AND_HOLIDAY_FILL_COLOR      
        ws.cell(row = curRow, column = curCol + 1).fill = WEEKEND_AND_HOLIDAY_FILL_COLOR      

        if(SMPspacerRow != curRow):
            ws.cell(row = curRow, column = curCol).value = "X"
            ws.cell(row = curRow, column = curCol + 1).value = "X"
            ws.cell(row = curRow, column = curCol).font = Font(name = 'Calibri', size = 15, bold = True)
            ws.cell(row = curRow, column = curCol + 1).font = Font(name = 'Calibri', size = 15, bold = True)
            SMPspacerRow  += 4
        curRow += 1

####################################################################
### Function Title: createDateCols()
### Arguments: worksheet, end column, date time object, start date,
### day date (date in the week number), holiday dates list
### Returns: the day date (day in the week number) so can be used for second sheet
### Description: creates/sets the width/borders/font/alignment for each cell
### in each date column. adjusts for if that date is a federal holiday or weekend
###################################################################
def createDateCols(ws, endCol, startDate, dayDate, holidayDates, weekendDates, noSMPCBOCs, smpCBOCs):
    #to get name of day (in number) from date
    # to get name of day from date
    dateNum = startDate
    dayName = calendar.day_abbr[dayDate]
    dayName = dayName.upper()

    # make start col and row 2 since dates start after CBOC col and header row
    curRow = 2
    curCol = 2
    regColWidth = 4.5 

    endNonSmpRows = NUM_FIXED_ROWS + (len(noSMPCBOCs) * 3) 
    endRow = endNonSmpRows + (len(smpCBOCs) * 4) - 1
    while (curCol <= (endCol * 2)):
        
        ws.column_dimensions[get_column_letter(curCol)].width = regColWidth 

        # set the day date and name for the date/name cells
        setDateInfo(ws, curCol, dayDate, dayName, dateNum, curRow)
        #set tech info (tech, arrival time)
        setTechInfo(ws, curCol, curRow)
        
        # increment to get to second part of each date column
        curCol += 1
        ws.column_dimensions[get_column_letter(curCol)].width = regColWidth

        # check if is weekend or holiday
        if(isWeekend(dayName) == True):
            weekendDates.append(dateNum)
            setFixedWeekendAndHolStyle(ws, curCol - 1, curRow, endNonSmpRows, endRow)
        elif(isHoliday(holidayDates, dateNum) == True):
            setFixedWeekendAndHolStyle(ws, curCol - 1, curRow, endNonSmpRows, endRow)

        # increment to get to first column of new date
        curCol += 1
        # increment the day and date
        dayDate += 1
        dateNum += 1
        #if day number is > 6, i.e. you've reached end of week, start week days over
        if (dayDate > 6):
            dayDate = 0
        dayName = calendar.day_abbr[dayDate]
        dayName = dayName.upper()

    return dayDate, weekendDates

####################################################################
### Function Title: getStartDate()
### Arguments: none
### Returns: start date (in format mm-dd-yyyy) and the year input by
### user (in format yyyy)
### Description: gets input from the user. checks if is a valid year
### in formay yyyy and if is greater than or equal to 2021 (current year).
### if input is invalid, prints error message to the user and keeps asking
### for input until valid input is received. will then use the user's input
### year to create a valid start date that can be used with strptime to get
### a datetime object (which must be in format mm-dd-yyyy). always starts with
### jan 1 of whatever year requested by the user since will create an entire
### year's worth of cboc sheets starting in january
####################################################################
def getStartDate():
    # get start date of the month from user
    while(True):
        userInputYear = input("\nEnter year in the format yyyy: ")
        if(isValidUserInput(userInputYear)):
            break
        else:
            print("Your entry was invalid or a previous year. Enter current year in the format yyyy:")

    # reformat start date input into string for datetime in format 01-01-20yy
    startDate = "01-01-" + userInputYear
    return startDate, userInputYear

####################################################################
### Function Title: createCBOCCOL()
### Arguments: worksheet
### Returns: nothing
### Description: creates the column/font/border for the cboc column
### (which has all cboc names and a row for frozens from each)
###################################################################
def createCBOCCol(ws):
    # create cboc col border and font
    
    # set cboc col width
    ws.column_dimensions[get_column_letter(CBOC_COL)].width = CBOC_COL_WIDTH
    
    # set cboc and date border and font
    ws.cell(row=2, column=CBOC_COL).font = TIMES_NEW_ROMAN_FONT
    ws.cell(row=2, column=CBOC_COL).border = CBOC_NAME_BORDER
    ws.cell(row=2, column=CBOC_COL).value = "CBOC/CORE"
    ws.cell(row=3, column=CBOC_COL).border = DATE_BORDER
    ws.cell(row=3, column=CBOC_COL).font = TIMES_NEW_ROMAN_FONT
    ws.cell(row=3, column=CBOC_COL).value = "Date"
    ws.cell(row=4, column=CBOC_COL).border = BIG_SPACE_BORDER

####################################################################
### Function Title: saveExcelFile()
### Arguments: current month, current year, the workbook, the date
### time object
### Returns: none 
### Description: creates new folder if none exists. changes to that
### directory either way. saves the current workbook to an excel file.
###if the month is a single digit (1 - 9), will add a 0 in front of
### it so the files alphebatize in the correct order in the folder.
####################################################################

def saveExcelFile(currMonth, currYear, wb, dateTimeObj):
    # if is jan, create folder for all the sheets. otherwise you're already in the folder
    if(currMonth == 1):
        # if directory doesn't exist already, create it. then change to that directory from base directory either way
        if(os.path.isdir(str(currYear) + CBOC) == False):
            os.mkdir(str(currYear) + CBOC)
        os.chdir(str(currYear) + CBOC)

    # save workbook to excel file
    if(currMonth < 10):
        strMonth = "0" + str(currMonth)
    else:
        strMonth = str(currMonth)
    wb.save(strMonth + calendar.month_name[dateTimeObj.month] + "_" + str(dateTimeObj.year) + CBOC + '.xlsx')  

####################################################################
### Function Title:
### Arguments:
### Returns:
### Description: 
###################################################################
def setFixedRowHeights(ws):
    ws.row_dimensions[HEADER_ROW].height = HEADER_ROW_HEIGHT
    ws.row_dimensions[CBOC_ROW].height = CBOC_ROW_HEIGHT 
    ws.row_dimensions[DATE_ROW].height = DATE_ROW_HEIGHT
    ws.row_dimensions[TECH_TOA_ROW].height = TECH_TOA_ROW_HEIGHT


####################################################################
### Function Title:
### Arguments:
### Returns:
### Description: 
###################################################################
def setVariableRowHeights(ws, nonSmpCbocs, SmpCbocs):
    curRow = CBOC_NAME_START_ROW
    endNonSmpRows = NUM_FIXED_ROWS + (len(nonSmpCbocs) * 3) 
    endRow = endNonSmpRows + (len(SmpCbocs) * 4) - 1

    # set row heights for non SMP cboc section
    while (curRow <= endNonSmpRows):
        # set height for cboc name portion
        ws.row_dimensions[curRow].height = CBOC_NAME_AND_FROZEN_ROW_HEIGHT 
        curRow += 1
        # set height for frozen portion
        ws.row_dimensions[curRow].height = CBOC_NAME_AND_FROZEN_ROW_HEIGHT
        curRow += 1
        # set height for spacer portion
        ws.row_dimensions[curRow].height = SPACER_ROW_HEIGHT
        curRow +=1

    # set row heights for SMP cboc section
    while (curRow <= endRow):
        # set height for cboc name portion
        ws.row_dimensions[curRow].height = CBOC_NAME_AND_FROZEN_ROW_HEIGHT
        curRow += 1
        # set height for frozen portion
        ws.row_dimensions[curRow].height = CBOC_NAME_AND_FROZEN_ROW_HEIGHT
        curRow += 1
        # set height for SMP portion
        ws.row_dimensions[curRow].height = SMP_ROW_HEIGHT
        curRow += 1
        # set height for spacer portion unless last row
        if(curRow < endRow):
            ws.row_dimensions[curRow].height = SPACER_ROW_HEIGHT
        curRow += 1

####################################################################
### Function Title:
### Arguments:
### Returns:
### Description: 
###################################################################
def createCBOCColBorders(ws, noSMPCBOCs, smpCBOCs):
    curRow = CBOC_NAME_START_ROW
    endNonSmpRows = NUM_FIXED_ROWS + (len(noSMPCBOCs) * 3) 
    endRow = endNonSmpRows + (len(smpCBOCs) * 4) - 1
    curCBOCIdx = 0
    
    # put in borders and values for non smp cbocs
    while(curRow < endNonSmpRows):
        # put in cboc name as value and border
        ws.cell(row = curRow, column = CBOC_COL).font = TIMES_NEW_ROMAN_FONT
        ws.cell(row = curRow, column = CBOC_COL).border = CBOC_NAME_ONLY_BORDER
        ws.cell(row = curRow, column = CBOC_COL).value = noSMPCBOCs[curCBOCIdx]
        curCBOCIdx += 1
        curRow += 1

        # put in frozen value and border
        ws.cell(row = curRow, column = CBOC_COL).font = TIMES_NEW_ROMAN_FONT
        ws.cell(row = curRow, column = CBOC_COL).border = FROZEN_ONLY_BORDER
        ws.cell(row = curRow, column = CBOC_COL).value = FZ
        curRow += 1

        # put in spacer border (unless last row)
        if(endNonSmpRows != endRow):
            ws.cell(row = curRow, column = CBOC_COL).border = SPACER_BORDER
        curRow += 1

    # put in borders and values for smp cbocs
    curCBOCIdx = 0
    while(curRow < endRow):
        # put in cboc name as value and border
        ws.cell(row = curRow, column = CBOC_COL).font = TIMES_NEW_ROMAN_FONT
        ws.cell(row = curRow, column = CBOC_COL).border = CBOC_NAME_ONLY_BORDER
        ws.cell(row = curRow, column = CBOC_COL).value = smpCBOCs[curCBOCIdx]
        curCBOCIdx += 1
        curRow += 1

        # put in frozen value and border
        ws.cell(row = curRow, column = CBOC_COL).font = TIMES_NEW_ROMAN_FONT
        ws.cell(row = curRow, column = CBOC_COL).border = FROZEN_SMP_BORDER
        ws.cell(row = curRow, column = CBOC_COL).value = FZ
        curRow += 1 

        # put in SMP value and border
        ws.cell(row = curRow, column = CBOC_COL).font = TIMES_NEW_ROMAN_FONT
        ws.cell(row = curRow, column = CBOC_COL).value = SMP
        ws.cell(row = curRow, column = CBOC_COL).border = SMP_ONLY_BORDER
        ws.cell(row = curRow, column = CBOC_COL).alignment = Alignment(horizontal='left',vertical='center')
        curRow += 1

        # put in spacer border (unless last row)
        if(curRow < endRow):
            ws.cell(row = curRow, column = CBOC_COL).border = SPACER_BORDER
        curRow += 1

####################################################################
### Function Title:
### Arguments:
### Returns:
### Description: 
###################################################################
def createBottomRowBorder(ws, noSMPCBOCs, smpCBOCs, endDate):
    endNonSmpRows = NUM_FIXED_ROWS + (len(noSMPCBOCs) * 3) 
    endRow = endNonSmpRows + (len(smpCBOCs) * 4) - 1
    endCol = endDate
    curCol = CBOC_COL

    while(curCol <= endCol):
        ws.cell(row = endRow + 1, column = curCol).border = BOTTOM_ROW_BORDER
        curCol += 1

####################################################################
### Function Title: main()
### Arguments: none
### Returns: none
### Description: has loop for each month of the given year chosen by
### the user. creates workbook and 2 worksheets per month. calls other
### functions used to get user info, get the datetime object, calculate
### the num of days in the month, calculate the federal holidays, set
### all row and column info for all cells in the cboc signin sheet,
### creates a folder for each year's worksheets (unless the folder
### already exists), and saves all excel spreadsheets using the month
### number/name formatted so that it will have them in correct month
### order in the folder (jan - dec for the given year).
###################################################################
def main():
    # create lists for both SMP and non SMP CBOC names
    smpCBOCs = []
    noSMPCBOCs = []

    # get CBOCs from file
    smpCBOCs, noSMPCBOCs = readFile.getCBOClists(smpCBOCs, noSMPCBOCs)
    
    currMonth = START_MONTH
    currDate, currYear = getStartDate()
    print("...excel spreadsheet creation in progress please wait...")
    
    while (currMonth <= END_MONTH):
        # create workbook (1st sheet at pos 0 created automatically)
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "1-15"
    
        # create 2nd sheet at pos 1
        ws2 = wb.create_sheet("16-End", 1)
        
        # Create date object in format mm-dd-yyyy from start date string
        dateTimeObj = datetime.strptime(currDate, "%m-%d-%Y")
        
        # get end date for the month
        endDate = calendar.monthrange(dateTimeObj.year, dateTimeObj.month)[1]
        
        # calculate holiday dates for the month
        holidayDates = []
        holidayDates = federalHolidayCalculator.calcFedHolidays(dateTimeObj) 

        #####################

        # set row heights for fixed portions of sheet
        setFixedRowHeights(ws1)
        setFixedRowHeights(ws2)

        # create cboc cell font/border/values for both sheets
        createCBOCCol(ws1)
        createCBOCCol(ws2)

        # create rest of cols (date cols) for both sheets
        dayDate = dateTimeObj.weekday()

        # update day date to be last day date from first sheet before passing to second sheet
        # also get weekend dates
        weekendDates = []
        dayDate, weekendDates = createDateCols(ws1, MID_DATE, 1, dayDate, holidayDates, weekendDates, noSMPCBOCs, smpCBOCs)
        dayDate, weekendDates = createDateCols(ws2, endDate - MID_DATE, MID_DATE + 1, dayDate, holidayDates, weekendDates, noSMPCBOCs, smpCBOCs)

        # create header for both sheets
        createHeader(ws1, HEADER_ROW, HEADER_AND_LABELS_COL, HEADER_ROW, (MID_DATE * 2) + 1, dateTimeObj)
        createHeader(ws2, HEADER_ROW, HEADER_AND_LABELS_COL, HEADER_ROW, ((endDate - MID_DATE) * 2) + 1, dateTimeObj)

        # set row heights for the variable portions of sheet
        setVariableRowHeights(ws1, noSMPCBOCs, smpCBOCs)
        setVariableRowHeights(ws2, noSMPCBOCs, smpCBOCs)

        # put in cboc name row values and borders
        createCBOCColBorders(ws1, noSMPCBOCs, smpCBOCs)
        createCBOCColBorders(ws2, noSMPCBOCs, smpCBOCs)

        #createSignatureBorders(ws1, MID_DATE * 2, noSMPCBOCs, smpCBOCs)
        #createSignatureBorders(ws2, (endDate - MID_DATE) * 2, noSMPCBOCs, smpCBOCs)

        # adjust bottom border of last row
        createBottomRowBorder(ws1, noSMPCBOCs, smpCBOCs, (MID_DATE * 2) + 1)
        createBottomRowBorder(ws2, noSMPCBOCs, smpCBOCs, ((endDate - MID_DATE) * 2) + 1)

        # put in grey fill in background of weekends/holidays
        #setVariableWeekendAndHolStyle(ws1, holidayDates, weekendDates, noSMPCBOCs, smpCBOCs, (MID_DATE * 2) + 1)
        #setVariableWeekendAndHolStyle(ws2, holidayDates, weekendDates, noSMPCBOCs, smpCBOCs, ((endDate - MID_DATE) * 2) + 1) 

        # save the current month's excel file
        saveExcelFile(currMonth, currYear, wb, dateTimeObj)
        
        #increment the month
        currMonth += 1
        currDate = str(currMonth) + "-01-" + str(currYear)
    
    # tell user what folder their files are in. wait for their input (enter) before exiting.
    print("\nSpreadsheets created, see folder: \"" + str(currYear) + "_cboc_signin_sheets\" for your files.")
    print("Press \"Enter\" to finish.")
    input()

if __name__ == "__main__":
    main()